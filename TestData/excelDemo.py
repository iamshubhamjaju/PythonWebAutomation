import openpyxl

book = openpyxl.load_workbook("C:\\Users\\sjaju\\OneDrive\\Desktop\\PythonDemo.xlsx")

sheet = book.active
Dict = {}
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# cell1 = sheet.cell(row=2, column=2).value = "Shubham"
# print(cell1)

# print(sheet.max_row)
# print(sheet.max_column)

# print(sheet['A5'].value)

for i in range(1, sheet.max_row):
    # if sheet.cell(row=i, column=1).value == "S.no":
    for j in range(2, sheet.max_column + 1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
