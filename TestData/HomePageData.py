import openpyxl

class HomePageData:
    test_HomePage_data = [{"Name": "Rahul", "Email": "yoshubhamjaju@gmail.com", "Password": "Juju", "Gender": "Female"},
                          {"Name": "Rahul", "Email": "yoshubhamjaju@gmail.com", "Password": "Juju", "Gender": "Male"}]

    # Dictionary [{"Key": "Value"}]
    @staticmethod
    def getTestData(testcase_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\sjaju\\OneDrive\\Desktop\\PythonDemo.xlsx")

        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]
